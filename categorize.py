#! python3
from lxml import html, etree
from pathlib import Path
import time
import os
import sys
import openpyxl
import re
import hashlib
import json
import requests

def search_colA(searching_for):
    counter = 2  # start on row after header
    while (counter < (ws.max_row + 1)):
        currentA = "A" + str(counter)
        valA = ws[currentA].value
        if valA == searching_for:
            found_on_counter = counter
            break
        counter = counter + 1
    return found_on_counter

class CurrentCell:
    def __init__(self, coord):
        self.coord = coord
        self.value = self.init_sheet_cellvalue()
    
    def init_sheet_cellvalue(self):
        valCell = str(ws[self.coord].value)
        return valCell
    
    def update_sheet_cellvalue(self):
        ws[self.coord].value = self.value

    def get_coord(self):
        return self.coord

    def get_value(self):
        return self.value

    def update_value(self, new_val):
        self.value = new_val

    def update_sheet(self):
        ws[self.coord].value = self.value



# Vars
# set me
subcatsDic = {
    "85" : 'faith-and-morals',
    "83" : 'education',
    "84" : 'family',
    "252" : 'history',
    "82" : 'biography'
}

# files and folders
data_folder = Path("C:/Users/chris/Desktop/migAssets")
source = "source/Content.xlsx"
file_to_open = data_folder / source
for k, v in subcatsDic.items():
    print("")
    save_as = v + "Content.xlsx"
    save_path = data_folder / save_as


    #Start
    #---------
    wb = openpyxl.load_workbook(file_to_open)
    ws = wb.active


    #get max row of excel sheet that contains data
    maxRow = None
    counter = 2  # start on row after header
    while (counter < (ws.max_row + 2)):
        currentA = "A" + str(counter)
        valA = ws[currentA].value
        if valA == None:
            maxRow = counter
            break
        counter = counter + 1

    if (maxRow == None):
        maxRow = counter - 1
        
    print("Max row: " + str(maxRow))
    print("")
    ids_toRemoveDic = {}

    # collect ids of rows to remove
    counter = 2 #start on row after header
    count_removes = 1
    while ( counter < maxRow ):  
        print("Processing row: " + str(counter))
        print("")
        # Create cell objects for current row
        currA_postid = CurrentCell( "A"+str(counter) )
        currB_cat = CurrentCell( "B"+str(counter) )

        # if empty 
        if currA_postid.get_value() == None:
            print("blank row! " + str(counter))
            sys.exit()

        elif currB_cat.get_value() != k:
            # add to Dictionary
            ids_toRemoveDic.update( {count_removes: currA_postid.get_value()} )
            count_removes += 1
        else:
            pass

        counter += 1
    
    print("count removes: " + count_removes)
    # After while loop, delete rows
    for key, value in ids_toRemoveDic.items():
        row_to_delete = search_colA(value)
        ws.delete_rows(row_to_delete)

    wb.save(save_path)
    wb.close()
    print('created: ' + str(save_path))
    print("")

print("Finished")

    