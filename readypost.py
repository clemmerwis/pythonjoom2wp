#! python3
from lxml import html, etree
from pathlib import Path
import time
import os
import sys
import openpyxl
import re
import hashlib
import json
import requests

def search_colA(searching_for):
    counter = 2  # start on row after header
    while (counter < (ws.max_row + 1)):
        currentA = "A" + str(counter)
        valA = ws[currentA].value
        if valA == searching_for:
            found_on_counter = counter
            break
        counter = counter + 1
    return found_on_counter


def get_image_url(val):
    image_tag = imageCheck(val)
    img_url_regex = re.compile(r'src="(.*(jpg|png))')
    img_url_mo = img_url_regex.search(image_tag)
    img_url = img_url_mo.group(1)
    return img_url
    

def imageCheck(val):
    check = ''
    img_check_regex = re.compile(r'<img.*(jpg|png)".*/>')
    img_check_mo = img_check_regex.search(val)
    if img_check_mo != None:
        check = img_check_mo.group()
    return check


def imgStart(val):
    img_start_regex = re.compile(r'<img.*src=\"images')
    img_start_mo = img_start_regex.search(val)
    if img_start_mo != None:
        img_start_mo = str(img_start_mo.group())
        val = val.replace(img_start_mo, '<img src="http://www.thenewamerican.com/images')
    return val


def imgEnd(val):
    img_end_regex = re.compile(r'((.jpg|png)\" .*)/>')
    img_end_mo = img_end_regex.search(val)
    if img_end_mo != None:
        img_end_mo = str(img_end_mo.group())
        if img_end_mo.find("jpg") != -1:
            val = val.replace(img_end_mo, '.jpg"/>')
        elif img_end_mo.find("png") != -1:
            val = val.replace(img_end_mo, '.png"/>')
    return val


def format_image(val):
    val = imgStart(val)
    val = imgEnd(val)
    return val



def get_youtube_embeds(val):
    embeds = list()
    embeds_check_regex = re.compile(r'(<p>|<p.*?)?({youtube}(.*){/youtube})(.*?p>)?')
    embeds_tuples = embeds_check_regex.findall(val)
    if embeds_tuples != None:
        for tup in embeds_tuples:
          embed = str(tup[2])
          embeds.append(embed)
    return embeds


def format_modulepos(val):
    mod_check_regex = re.compile(r'(<p>|<p.*?)({modulepos .*})(.*?p>)')
    mod_tuples = mod_check_regex.findall(val)
    if mod_tuples != None:
        for tup in mod_tuples:
            toFind1 = str(tup[0])
            toFind2 = str(tup[1])
            toFind3 = str(tup[2])
            toFind = toFind1 + toFind2 + toFind3
            val = val.replace(toFind, "")
    return val


def create_video_iframe(videoID):
    iframe = f'<iframe frameborder="0" allowfullscreen="allowfullscreen" src="https://www.youtube.com/embed/{videoID}?rel=0&amp;showinfo=0&amp;autoplay=0"></iframe>'
    return iframe


def format_ytvideos(val):
    embeds = get_youtube_embeds(val)
    yt_check_regex = re.compile(r'(<p>|<p.*?)?({youtube}.*{/youtube})(.*?p>)?')
    yt_tuples = yt_check_regex.findall(val)
    if yt_tuples != None:
        for tup in yt_tuples:
            for embed in embeds:
                if embed in str(tup[1]): 
                    yt_iframe = create_video_iframe(embed)
                    toFind1 = str(tup[0])
                    toFind2 = str(tup[1])
                    toFind3 = str(tup[2])
                    toFind = toFind1 + toFind2 + toFind3
                    val = val.replace(toFind, yt_iframe)
    return val


def insert_before_first_p_tag(val, to_insert):
    append_portion = val.split('<p')[0]
    val = val.replace( append_portion, to_insert, 1 )
    return val


def highslide_format(val):
    slide_check_regex = re.compile(r'<p.*?\"highslide\".*?>')
    slide_check_mo = slide_check_regex.search(val)
    if slide_check_mo != None:
        toFind = slide_check_mo.group()
        val = val.replace(toFind, "")
    return val


def getHash(postid):
    the_hash = hashlib.md5(("Image" + str(postid)).encode('utf-8')).hexdigest()
    return the_hash


class CurrentCell:
    def __init__(self, coord):
        self.coord = coord
        self.value = self.init_sheet_cellvalue()
    
    def init_sheet_cellvalue(self):
        valCell = str(ws[self.coord].value)
        return valCell
    
    def update_sheet_cellvalue(self):
        ws[self.coord].value = self.value

    def get_coord(self):
        return self.coord

    def get_value(self):
        return self.value

    def update_value(self, new_val):
        self.value = new_val

    def update_sheet(self):
        ws[self.coord].value = self.value



# Vars
# change me
# Check for 3 arguments 
if sys.argv[3] == None:
    print('"Script requires arguments: "catnum", "catname", "subtcat"')
    sys.exit()
catnum = sys.argv[1]
catname = sys.argv[2]
subcat = sys.argv[3]
# cat = "83"
catdomain = catname + '/' + subcat + '/' + 'item/'
joomla_domain = 'https://www.thenewamerican.com/' 

# files and folders
jsondata_folder = Path("C:/Users/chris/Desktop/migAssets/json")
data_folder = Path("C:/Users/chris/Desktop/migAssets")
file_name = catnum + "Content.xlsx"
file_to_open = data_folder / file_name
save_as = catnum + "Content-formatted.xlsx"
save_path = data_folder / save_as


#Start
#---------
wb = openpyxl.load_workbook(file_to_open)
ws = wb.active


#get max row of excel sheet that contains data
maxRow = None
counter = 2  # start on row after header
while (counter < (ws.max_row + 2)):
    currentA = "A" + str(counter)
    valA = ws[currentA].value
    if valA == None:
        maxRow = counter
        break
    counter = counter + 1

if (maxRow == None):
    maxRow = counter - 1
    
print("Max row: " + str(maxRow))

idsUrls_Dic = {}
idsSkipped_Dic = {}
# Format spreadsheet
counter = 2 #start on row after header
while ( counter < maxRow ):  
    # Create cell objects for current row
    currA_postid = CurrentCell( "A"+str(counter) )
    currG_excerpt = CurrentCell( "G"+str(counter) )
    currE_content = CurrentCell( "E"+str(counter) )
    currK_postname = CurrentCell( "K"+str(counter) )

    # if empty 
    if currA_postid.get_value() == None:
        print(str(counter))
        sys.exit()

    # if is a Video post
    if (currG_excerpt.get_value().find("VIDEO -") ) != -1:
        # for now just keep track of video posts
        print(currA_postid.get_value())
        # add to skipped json
        idsSkipped_Dic.update( {"row" + str(counter): currA_postid.get_value()} )
        counter += 1
        continue
    # else is a normal post
    else:
        # remove anything before first p tag
        currG_excerpt.update_value( insert_before_first_p_tag(currG_excerpt.get_value(), "") )

        # format p class obscuring image on some & update value
        currG_excerpt.update_value( highslide_format(currG_excerpt.get_value()) )

        # If no image, get image from ID and hash
        if imageCheck(currG_excerpt.get_value()) == '':
            image_hash = getHash(currG_excerpt.get_value())
            the_image_url = 'https://www.thenewamerican.com/media/k2/items/src/' + image_hash + '.jpg'

            # check if url exists
            r = requests.get(the_image_url)
            if r.status_code == 404:
                webpageLink = joomla_domain + catdomain + currA_postid.get_value() +  '-' + currK_postname.get_value()
                page = requests.get(webpageLink)

                # convert the data received into searchable HTML
                extractedHtml = html.fromstring(page.content)

                # use an XPath query to find the image link (the 'src' attribute of the 'img' tag).
                imageSrc = extractedHtml.xpath("//img/@src")
                for src in imageSrc:
                    jpeg_found = False
                    if ".jpg" in src:
                        jpeg_found = True
                        src = src[1:]
                        the_image_url = joomla_domain + src
                        break
                    if ".png" in src:
                        jpeg_found = True
                        src = src[1:]
                        the_image_url = joomla_domain + src
                        break
            
            # add to Dictionary of ids and urls
            idsUrls_Dic.update( {currA_postid.get_value(): the_image_url} )

            # add image url to column R
            ws["Q"+str(counter)].value = idsUrls_Dic[currA_postid.get_value()]
        else:
            # format image in column G & update value
            currG_excerpt.update_value( format_image(currG_excerpt.get_value()) )
            
            # get the image url
            the_image_url = get_image_url(currG_excerpt.get_value())

            # add to Dictionary of ids and urls
            idsUrls_Dic.update( {currA_postid.get_value(): the_image_url} )

            # add image url to column R
            ws["Q"+str(counter)].value = idsUrls_Dic[currA_postid.get_value()]

            # remove image from excerpt
            currG_excerpt.update_value( currG_excerpt.get_value().replace( imageCheck(currG_excerpt.get_value()), "", 1 ) ) 

    #update post_content to excerpt + content
    currE_content.update_value( currG_excerpt.get_value() + currE_content.get_value() )

    # remove revive ads
    currE_content.update_value( format_modulepos(currE_content.get_value()) )

    # replace embed codes with youtube iframes
    currE_content.update_value( format_ytvideos(currE_content.get_value()) )
        
    #concat id to post_name
    currK_postname.update_value( currA_postid.get_value() + '-' + currK_postname.get_value() )
    
    # update cell values for current row
    currA_postid.update_sheet()
    currG_excerpt.update_sheet()
    currE_content.update_sheet()
    currK_postname.update_sheet()

    counter += 1
        
# After while loop, delete rows
for key, value in idsSkipped_Dic.items():
    row_to_delete = search_colA(value)
    ws.delete_rows(row_to_delete)

# Writing to Skipped.json 
jsonified = json.dumps(idsSkipped_Dic, indent = 4)
fname = catnum + "skipped.json"
json_file = jsondata_folder / fname
with open(json_file, "w+") as outfile: 
    outfile.write(jsonified)
    outfile.close() 
print('created: ' + str(json_file))

# Writing to Feats.json 
jsonified = json.dumps(idsUrls_Dic, indent = 4)
fname = catnum + "feats.json"
json_file = jsondata_folder / fname
with open(json_file, "w+") as outfile: 
    outfile.write(jsonified) 
    outfile.close()
print('created: ' + str(json_file))


wb.save(save_path)
print('created: ' + str(save_path))

print("Finished")

    